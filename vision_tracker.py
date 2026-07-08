from __future__ import annotations

import sqlite3
import sys
import re
from contextlib import closing
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


APP_TITLE = "Vision Issue Tracker"
APP_DIR = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path.cwd()
DB_PATH = APP_DIR / "data" / "vision_issues.db"

LINES = ["1-1", "1-2", "2-1", "2-2"]
INSTRUMENTS = [
    "Pinhole",
    "Pouch Align",
    "Lead",
    "Sealing",
    "Lead Align",
    "Welding(+)",
    "Welding(-)",
]
INSTRUMENT_SEPARATOR = " / "
WORKERS = ["Hojun Kwak", "Kijung Kim", "Jihoon Yun", "Jisub Yun"]
ACTIVE_STATUS_OPTIONS = ["Action Required", "Monitoring"]
STATUS_OPTIONS = ACTIVE_STATUS_OPTIONS + ["Resolved"]
CATEGORY_MAP = {
    "Hardware": ["Camera", "Lighting"],
    "Software": ["Program Crash", "Program Update", "UI", "PLC", "Other"],
    "Recipe": ["Overkill", "Underkill", "Add Measure", "Bypass/Unbypass"],
    "Camera Grab Fail": [""],
    "Production": [""],
    "Other": [""],
}
CATEGORIES = list(CATEGORY_MAP.keys())
VERSION_GROUPS = {
    "Welding": ["Welding(+)", "Welding(-)"],
    "Common": ["Pinhole", "Pouch Align", "Lead Align"],
    "New Lead": ["Lead"],
    "Sealing": ["Sealing"],
}
INSTRUMENT_GROUP = {
    instrument: group_name
    for group_name, instruments in VERSION_GROUPS.items()
    for instrument in instruments
}
NO_ALGO_INSTRUMENTS = {"Sealing"}


def instrument_uses_algo(instrument: str) -> bool:
    return instrument not in NO_ALGO_INSTRUMENTS


def version_group_uses_algo(group_name: str) -> bool:
    return any(instrument_uses_algo(instrument) for instrument in VERSION_GROUPS.get(group_name, []))


SW_VERSION_PATTERN = re.compile(r"^\s*(\d{6})\.(\d{4})\s*$")
ALGO_VERSION_PATTERN = re.compile(r"^\s*(\d+)\.(\d+)\.(\d+)\.(\d+)\s*$")


def sw_version_sort_key(value: str) -> tuple[int, int] | None:
    match = SW_VERSION_PATTERN.match(value or "")
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def algo_version_sort_key(value: str) -> tuple[int, int, int, int] | None:
    match = ALGO_VERSION_PATTERN.match(value or "")
    if not match:
        return None
    return tuple(int(part) for part in match.groups())


def version_sort_key(value: str, component: str | None = None) -> tuple[int, ...] | None:
    if component == "sw":
        key = sw_version_sort_key(value)
        if key is not None:
            return key
    elif component == "algo":
        key = algo_version_sort_key(value)
        if key is not None:
            return key
    else:
        key = sw_version_sort_key(value) or algo_version_sort_key(value)
        if key is not None:
            return key

    digits: list[str] = []
    current = ""
    for char in value:
        if char.isdigit():
            current += char
        elif current:
            digits.append(current)
            current = ""
    if current:
        digits.append(current)
    if not digits:
        return None
    return tuple(int(part) for part in digits)


def split_instruments(value: str) -> list[str]:
    instruments = [item.strip() for item in value.split("/") if item.strip()]
    return instruments


def format_instruments(values: list[str] | tuple[str, ...] | set[str]) -> str:
    ordered = [instrument for instrument in INSTRUMENTS if instrument in values]
    return INSTRUMENT_SEPARATOR.join(ordered)


@dataclass(frozen=True)
class IssueInput:
    issue_time: str
    line: str
    instrument: str
    worker: str
    category: str
    subcategory: str
    title: str
    description: str
    status: str = ACTIVE_STATUS_OPTIONS[0]
    resolved_time: str = ""
    resolution_notes: str = ""


@dataclass(frozen=True)
class VersionInput:
    update_time: str
    group_name: str
    line: str
    instrument: str
    sw_version: str
    algo_version: str
    description: str
    worker: str
    sw_description: str = ""
    algo_description: str = ""


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def downtime_duration(issue_time: str, end_time: datetime | None = None) -> str:
    end_time = end_time or datetime.now()
    try:
        start_time = datetime.strptime(issue_time, "%Y-%m-%d %H:%M")
    except ValueError:
        return ""
    minutes = max(0, int((end_time - start_time).total_seconds() // 60))
    hours, remaining_minutes = divmod(minutes, 60)
    return f"{hours:02d}:{remaining_minutes:02d}"


def clean_source_metadata(value: str | None) -> str:
    if not value:
        return ""
    metadata_patterns = [
        re.compile(r"^\s*Source\s+No\.?\s*:.*$", re.IGNORECASE),
        re.compile(r"^\s*Original\s+Vision\s*:.*$", re.IGNORECASE),
    ]
    kept_lines = [
        line
        for line in str(value).splitlines()
        if not any(pattern.match(line) for pattern in metadata_patterns)
    ]
    cleaned = "\n".join(kept_lines).strip()
    return re.sub(r"\n{3,}", "\n\n", cleaned)


def connect(db_path: Path = DB_PATH) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_column(conn: sqlite3.Connection, table_name: str, column_name: str, definition: str) -> None:
    columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table_name})")}
    if column_name not in columns:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}")


def clean_issue_source_metadata(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT id, description, resolution_notes
        FROM issues
        WHERE COALESCE(description, '') LIKE '%Source No%'
           OR COALESCE(description, '') LIKE '%Original Vision%'
           OR COALESCE(resolution_notes, '') LIKE '%Source No%'
           OR COALESCE(resolution_notes, '') LIKE '%Original Vision%'
        """
    ).fetchall()
    for row in rows:
        description = clean_source_metadata(row["description"])
        resolution_notes = clean_source_metadata(row["resolution_notes"])
        if description != (row["description"] or "") or resolution_notes != (row["resolution_notes"] or ""):
            conn.execute(
                """
                UPDATE issues
                SET description = ?, resolution_notes = ?
                WHERE id = ?
                """,
                (description, resolution_notes, row["id"]),
            )


def initialize_database(db_path: Path = DB_PATH) -> None:
    with closing(connect(db_path)) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS issues (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                issue_time TEXT NOT NULL,
                resolved_time TEXT,
                line TEXT NOT NULL,
                instrument TEXT NOT NULL,
                worker TEXT NOT NULL,
                category TEXT NOT NULL,
                subcategory TEXT,
                title TEXT NOT NULL,
                description TEXT,
                status TEXT NOT NULL,
                resolution_notes TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_issues_lookup
            ON issues(status, category, subcategory, line, instrument, issue_time)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS version_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_name TEXT NOT NULL,
                sw_version TEXT NOT NULL,
                algo_version TEXT NOT NULL,
                description TEXT,
                worker TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_version_templates_lookup
            ON version_templates(group_name, updated_at)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS version_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                update_time TEXT NOT NULL,
                group_name TEXT NOT NULL,
                line TEXT NOT NULL,
                instrument TEXT NOT NULL,
                sw_version TEXT NOT NULL,
                algo_version TEXT NOT NULL,
                description TEXT,
                worker TEXT NOT NULL,
                created_issue_id INTEGER,
                FOREIGN KEY(created_issue_id) REFERENCES issues(id)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_version_history_lookup
            ON version_history(line, instrument, group_name, update_time)
            """
        )
        ensure_column(conn, "version_templates", "sw_description", "TEXT")
        ensure_column(conn, "version_templates", "algo_description", "TEXT")
        ensure_column(conn, "version_history", "sw_description", "TEXT")
        ensure_column(conn, "version_history", "algo_description", "TEXT")
        ensure_column(conn, "version_history", "sw_touched", "INTEGER")
        ensure_column(conn, "version_history", "algo_touched", "INTEGER")
        backfill_version_history_component_flags(conn)
        backfill_version_history_descriptions(conn)
        conn.execute("UPDATE issues SET status = 'Action Required' WHERE status = 'Open'")
        conn.execute("UPDATE issues SET status = 'Monitoring' WHERE status = 'In Progress'")
        conn.execute("UPDATE issues SET subcategory = 'Overkill' WHERE subcategory = 'Overkill(False Reject)'")
        conn.execute("UPDATE issues SET subcategory = 'Underkill' WHERE subcategory = 'Underkill(False Accept)'")
        clean_issue_source_metadata(conn)
        conn.commit()


def validate_issue(issue: IssueInput) -> list[str]:
    errors: list[str] = []
    required = {
        "Issue time": issue.issue_time,
        "Line": issue.line,
        "Instrument": issue.instrument,
        "Worker": issue.worker,
        "Category": issue.category,
        "Title": issue.title,
        "Status": issue.status,
    }
    for label, value in required.items():
        if not value.strip():
            errors.append(f"{label} is required.")
    try:
        datetime.strptime(issue.issue_time, "%Y-%m-%d %H:%M")
    except ValueError:
        errors.append("Issue time must use YYYY-MM-DD HH:MM format.")
    if issue.line not in LINES:
        errors.append("Line is not valid.")
    issue_instruments = split_instruments(issue.instrument)
    if not issue_instruments:
        errors.append("Instrument is required.")
    invalid_instruments = [instrument for instrument in issue_instruments if instrument not in INSTRUMENTS]
    if invalid_instruments:
        errors.append("Instrument is not valid.")
    if issue.category not in CATEGORY_MAP:
        errors.append("Category is not valid.")
    if issue.status not in STATUS_OPTIONS:
        errors.append("Status is not valid.")
    allowed_subcategories = CATEGORY_MAP.get(issue.category, [])
    if issue.subcategory and issue.subcategory not in allowed_subcategories:
        errors.append("Subcategory is not valid for the selected category.")
    return errors


def create_issue(issue: IssueInput, db_path: Path = DB_PATH) -> int:
    errors = validate_issue(issue)
    if errors:
        raise ValueError("\n".join(errors))

    with closing(connect(db_path)) as conn:
        cursor = conn.execute(
            """
            INSERT INTO issues (
                created_at, issue_time, resolved_time, line, instrument, worker,
                category, subcategory, title, description, status, resolution_notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                now_text(),
                issue.issue_time,
                issue.resolved_time,
                issue.line,
                issue.instrument,
                issue.worker,
                issue.category,
                issue.subcategory,
                issue.title,
                issue.description,
                issue.status,
                issue.resolution_notes,
            ),
        )
        conn.commit()
        return int(cursor.lastrowid)


def update_issue(issue_id: int, issue: IssueInput, db_path: Path = DB_PATH) -> None:
    errors = validate_issue(issue)
    if errors:
        raise ValueError("\n".join(errors))

    with closing(connect(db_path)) as conn:
        conn.execute(
            """
            UPDATE issues
            SET issue_time = ?, resolved_time = ?, line = ?, instrument = ?,
                worker = ?, category = ?, subcategory = ?, title = ?,
                description = ?, status = ?, resolution_notes = ?
            WHERE id = ?
            """,
            (
                issue.issue_time,
                issue.resolved_time,
                issue.line,
                issue.instrument,
                issue.worker,
                issue.category,
                issue.subcategory,
                issue.title,
                issue.description,
                issue.status,
                issue.resolution_notes,
                issue_id,
            ),
        )
        conn.commit()


def resolve_issue(issue_id: int, notes: str = "", db_path: Path = DB_PATH) -> None:
    with closing(connect(db_path)) as conn:
        row = conn.execute("SELECT issue_time FROM issues WHERE id = ?", (issue_id,)).fetchone()
        duration = downtime_duration(row["issue_time"]) if row else ""
        conn.execute(
            """
            UPDATE issues
            SET status = 'Resolved',
                resolved_time = COALESCE(NULLIF(resolved_time, ''), ?),
                resolution_notes = ?
            WHERE id = ?
            """,
            (duration, notes, issue_id),
        )
        conn.commit()


def set_issue_status(issue_id: int, status: str, db_path: Path = DB_PATH) -> None:
    if status not in STATUS_OPTIONS:
        raise ValueError("Status is not valid.")
    with closing(connect(db_path)) as conn:
        if status == "Resolved":
            row = conn.execute("SELECT issue_time FROM issues WHERE id = ?", (issue_id,)).fetchone()
            duration = downtime_duration(row["issue_time"]) if row else ""
            conn.execute(
                """
                UPDATE issues
                SET status = ?,
                    resolved_time = COALESCE(NULLIF(resolved_time, ''), ?)
                WHERE id = ?
                """,
                (status, duration, issue_id),
            )
        else:
            conn.execute("UPDATE issues SET status = ? WHERE id = ?", (status, issue_id))
        conn.commit()


def delete_issue(issue_id: int, db_path: Path = DB_PATH) -> None:
    with closing(connect(db_path)) as conn:
        conn.execute("DELETE FROM issues WHERE id = ?", (issue_id,))
        conn.commit()


def validate_version_update(version: VersionInput) -> list[str]:
    errors: list[str] = []
    required = {
        "Update time": version.update_time,
        "Group": version.group_name,
        "Line": version.line,
        "Instrument": version.instrument,
        "SW Version": version.sw_version,
        "Worker": version.worker,
    }
    if instrument_uses_algo(version.instrument):
        required["Algo Version"] = version.algo_version
    for label, value in required.items():
        if not value.strip():
            errors.append(f"{label} is required.")
    try:
        datetime.strptime(version.update_time, "%Y-%m-%d %H:%M")
    except ValueError:
        errors.append("Update time must use YYYY-MM-DD HH:MM format.")
    if version.group_name not in VERSION_GROUPS:
        errors.append("Version group is not valid.")
    if version.line not in LINES:
        errors.append("Line is not valid.")
    if version.instrument not in INSTRUMENTS:
        errors.append("Instrument is not valid.")
    if version.instrument and version.group_name != INSTRUMENT_GROUP.get(version.instrument):
        errors.append("Instrument is not part of the selected version group.")
    return errors


def split_version_description(description: str | None) -> tuple[str, str]:
    value = (description or "").strip()
    sw_marker = "[SW Description]"
    algo_marker = "[Algo Description]"
    if sw_marker in value or algo_marker in value:
        sw_text = value
        algo_text = ""
        if sw_marker in value:
            sw_text = value.split(sw_marker, 1)[1]
        if algo_marker in sw_text:
            sw_text, algo_text = sw_text.split(algo_marker, 1)
        elif algo_marker in value:
            _, algo_text = value.split(algo_marker, 1)
        return sw_text.strip(), algo_text.strip()
    return value, ""


def combine_version_description(sw_description: str, algo_description: str, uses_algo: bool) -> str:
    sw_text = sw_description.strip()
    algo_text = algo_description.strip()
    if not uses_algo:
        return sw_text
    parts = []
    if sw_text:
        parts.append(f"[SW Description]\n{sw_text}")
    if algo_text:
        parts.append(f"[Algo Description]\n{algo_text}")
    return "\n\n".join(parts)


def version_description_parts(row: sqlite3.Row) -> tuple[str, str]:
    keys = set(row.keys())
    fallback_sw, fallback_algo = split_version_description(row["description"] if "description" in keys else "")
    sw_description = (row["sw_description"] or "").strip() if "sw_description" in keys else ""
    algo_description = (row["algo_description"] or "").strip() if "algo_description" in keys else ""
    if not sw_description:
        sw_description = fallback_sw
    if not algo_description:
        algo_description = fallback_algo
    if sw_description or algo_description:
        return sw_description, algo_description
    return "", ""


def infer_version_history_component_flags(row: sqlite3.Row) -> tuple[bool, bool]:
    sw_description, algo_description = version_description_parts(row)
    uses_algo = True
    keys = set(row.keys())
    if "instrument" in keys:
        uses_algo = instrument_uses_algo(row["instrument"])
    elif "group_name" in keys:
        uses_algo = version_group_uses_algo(row["group_name"])

    if not uses_algo:
        return True, False
    if sw_description and not algo_description:
        return True, False
    if algo_description and not sw_description:
        return False, True
    return True, True


def backfill_version_history_component_flags(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT id, group_name, instrument, description, sw_description, algo_description,
               sw_touched, algo_touched
        FROM version_history
        WHERE sw_touched IS NULL OR algo_touched IS NULL
        """
    ).fetchall()
    for row in rows:
        sw_touched, algo_touched = infer_version_history_component_flags(row)
        conn.execute(
            """
            UPDATE version_history
            SET sw_touched = ?, algo_touched = ?
            WHERE id = ?
            """,
            (1 if sw_touched else 0, 1 if algo_touched else 0, row["id"]),
        )


def version_component_description_from_templates(
    conn: sqlite3.Connection,
    group_name: str,
    component: str,
    version: str,
) -> str:
    if not version.strip():
        return ""
    version_column = "sw_version" if component == "sw" else "algo_version"
    rows = conn.execute(
        f"""
        SELECT description, sw_description, algo_description
        FROM version_templates
        WHERE group_name = ? AND {version_column} = ?
        ORDER BY updated_at DESC, id DESC
        """,
        (group_name, version.strip()),
    ).fetchall()
    for row in rows:
        sw_description, algo_description = version_description_parts(row)
        description = sw_description if component == "sw" else algo_description
        if description:
            return description
    return ""


def backfill_version_history_descriptions(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT id, group_name, instrument, sw_version, algo_version, description,
               sw_description, algo_description, sw_touched, algo_touched
        FROM version_history
        WHERE COALESCE(description, '') = ''
           OR COALESCE(sw_description, '') = ''
           OR COALESCE(algo_description, '') = ''
        """
    ).fetchall()
    for row in rows:
        sw_description, algo_description = version_description_parts(row)
        sw_touched, algo_touched = version_history_component_flags(row)
        if sw_touched and not sw_description:
            sw_description = version_component_description_from_templates(
                conn, row["group_name"], "sw", row["sw_version"]
            )
        if algo_touched and not algo_description:
            algo_description = version_component_description_from_templates(
                conn, row["group_name"], "algo", row["algo_version"]
            )
        if not version_group_uses_algo(row["group_name"]):
            algo_description = ""
        description = combine_version_description(
            sw_description,
            algo_description,
            version_group_uses_algo(row["group_name"]),
        )
        if (
            description != (row["description"] or "")
            or sw_description != (row["sw_description"] or "")
            or algo_description != (row["algo_description"] or "")
        ):
            conn.execute(
                """
                UPDATE version_history
                SET description = ?, sw_description = ?, algo_description = ?
                WHERE id = ?
                """,
                (description, sw_description, algo_description, row["id"]),
            )


def refresh_combined_version_descriptions(conn: sqlite3.Connection, table_name: str, group_name: str) -> None:
    rows = conn.execute(
        f"""
        SELECT id, description, sw_description, algo_description
        FROM {table_name}
        WHERE group_name = ?
        """,
        (group_name,),
    ).fetchall()
    uses_algo = version_group_uses_algo(group_name)
    for row in rows:
        sw_description, algo_description = version_description_parts(row)
        combined = combine_version_description(sw_description, algo_description, uses_algo)
        conn.execute(
            f"UPDATE {table_name} SET description = ? WHERE id = ?",
            (combined, row["id"]),
        )


def save_version_template(
    group_name: str,
    sw_version: str,
    algo_version: str,
    description: str,
    worker: str,
    db_path: Path = DB_PATH,
    sw_description: str | None = None,
    algo_description: str | None = None,
) -> int:
    if group_name not in VERSION_GROUPS:
        raise ValueError("Version group is not valid.")
    if not sw_version.strip():
        raise ValueError("SW Version is required.")
    if version_group_uses_algo(group_name) and not algo_version.strip():
        raise ValueError("Algo Version is required.")
    if sw_description is None and algo_description is None:
        sw_description_value, algo_description_value = split_version_description(description)
    else:
        sw_description_value = (sw_description or "").strip()
        algo_description_value = (algo_description or "").strip()
    if not version_group_uses_algo(group_name):
        algo_description_value = ""
    timestamp = now_text()
    with closing(connect(db_path)) as conn:
        if not sw_description_value:
            sw_description_value = version_component_description_from_templates(
                conn, group_name, "sw", sw_version
            )
        if version_group_uses_algo(group_name) and not algo_description_value:
            algo_description_value = version_component_description_from_templates(
                conn, group_name, "algo", algo_version
            )
        combined_description = (
            combine_version_description(
                sw_description_value,
                algo_description_value,
                version_group_uses_algo(group_name),
            )
            or description
        )
        existing = conn.execute(
            """
            SELECT id, description, sw_description, algo_description FROM version_templates
            WHERE group_name = ? AND sw_version = ? AND algo_version = ?
            ORDER BY id DESC LIMIT 1
            """,
            (group_name, sw_version.strip(), algo_version.strip()),
        ).fetchone()
        if existing:
            existing_sw_description, existing_algo_description = version_description_parts(existing)
            if not sw_description_value:
                sw_description_value = existing_sw_description
            if version_group_uses_algo(group_name) and not algo_description_value:
                algo_description_value = existing_algo_description
            combined_description = (
                combine_version_description(
                    sw_description_value,
                    algo_description_value,
                    version_group_uses_algo(group_name),
                )
                or existing["description"]
                or description
            )
            conn.execute(
                """
                UPDATE version_templates
                SET description = ?, sw_description = ?, algo_description = ?, worker = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    combined_description,
                    sw_description_value,
                    algo_description_value,
                    worker,
                    timestamp,
                    existing["id"],
                ),
            )
            conn.commit()
            return int(existing["id"])
        cursor = conn.execute(
            """
            INSERT INTO version_templates (
                group_name, sw_version, algo_version, description, sw_description, algo_description,
                worker, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                group_name,
                sw_version.strip(),
                algo_version.strip(),
                combined_description,
                sw_description_value,
                algo_description_value,
                worker,
                timestamp,
                timestamp,
            ),
        )
        conn.commit()
        return int(cursor.lastrowid)


def recent_version_templates(group_name: str, limit: int = 3, db_path: Path = DB_PATH) -> list[sqlite3.Row]:
    with closing(connect(db_path)) as conn:
        return list(
            conn.execute(
                """
                SELECT id, group_name, sw_version, algo_version, description, sw_description,
                       algo_description, worker, created_at, updated_at
                FROM version_templates
                WHERE group_name = ?
                ORDER BY updated_at DESC, id DESC
                LIMIT ?
                """,
                (group_name, limit),
            )
        )


def get_version_template(template_id: int, db_path: Path = DB_PATH) -> sqlite3.Row | None:
    with closing(connect(db_path)) as conn:
        return conn.execute(
            """
            SELECT id, group_name, sw_version, algo_version, description, sw_description,
                   algo_description, worker, created_at, updated_at
            FROM version_templates
            WHERE id = ?
            """,
            (template_id,),
        ).fetchone()


def version_component_templates(
    group_name: str,
    component: str,
    limit: int = 50,
    db_path: Path = DB_PATH,
) -> list[dict[str, str]]:
    if group_name not in VERSION_GROUPS:
        raise ValueError("Version group is not valid.")
    if component not in {"sw", "algo"}:
        raise ValueError("Version component is not valid.")
    if component == "algo" and not version_group_uses_algo(group_name):
        return []

    version_column = "sw_version" if component == "sw" else "algo_version"
    with closing(connect(db_path)) as conn:
        rows = conn.execute(
            """
            SELECT id, group_name, sw_version, algo_version, description, sw_description,
                   algo_description, worker, created_at, updated_at
            FROM version_templates
            WHERE group_name = ?
            ORDER BY updated_at DESC, id DESC
            """,
            (group_name,),
        ).fetchall()

    templates: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in rows:
        version = (row[version_column] or "").strip()
        if not version or version in seen:
            continue
        sw_description, algo_description = version_description_parts(row)
        templates.append(
            {
                "id": str(row["id"]),
                "group_name": row["group_name"],
                "component": component,
                "version": version,
                "description": sw_description if component == "sw" else algo_description,
                "worker": row["worker"],
                "created_at": row["created_at"],
                "updated_at": row["updated_at"],
            }
        )
        seen.add(version)
        if len(templates) >= limit:
            break
    return templates


def update_version_component_template(
    group_name: str,
    component: str,
    old_version: str,
    new_version: str,
    description: str,
    worker: str,
    db_path: Path = DB_PATH,
) -> None:
    if group_name not in VERSION_GROUPS:
        raise ValueError("Version group is not valid.")
    if component not in {"sw", "algo"}:
        raise ValueError("Version component is not valid.")
    if component == "algo" and not version_group_uses_algo(group_name):
        raise ValueError("This version group does not use Algo versions.")
    if not old_version.strip():
        raise ValueError("Select a version first.")
    if not new_version.strip():
        raise ValueError("Version is required.")

    version_column = "sw_version" if component == "sw" else "algo_version"
    description_column = "sw_description" if component == "sw" else "algo_description"
    timestamp = now_text()
    with closing(connect(db_path)) as conn:
        conn.execute(
            f"""
            UPDATE version_templates
            SET {version_column} = ?, {description_column} = ?, worker = ?, updated_at = ?
            WHERE group_name = ? AND {version_column} = ?
            """,
            (new_version.strip(), description, worker, timestamp, group_name, old_version.strip()),
        )
        conn.execute(
            f"""
            UPDATE version_history
            SET {version_column} = ?, {description_column} = ?, worker = ?
            WHERE group_name = ? AND {version_column} = ?
            """,
            (new_version.strip(), description, worker, group_name, old_version.strip()),
        )
        refresh_combined_version_descriptions(conn, "version_templates", group_name)
        refresh_combined_version_descriptions(conn, "version_history", group_name)
        conn.commit()


def delete_version_component_template(
    group_name: str,
    component: str,
    version: str,
    db_path: Path = DB_PATH,
) -> None:
    if group_name not in VERSION_GROUPS:
        raise ValueError("Version group is not valid.")
    if component not in {"sw", "algo"}:
        raise ValueError("Version component is not valid.")
    if not version.strip():
        return
    version_column = "sw_version" if component == "sw" else "algo_version"
    with closing(connect(db_path)) as conn:
        conn.execute(
            f"DELETE FROM version_templates WHERE group_name = ? AND {version_column} = ?",
            (group_name, version.strip()),
        )
        conn.execute(
            f"DELETE FROM version_history WHERE group_name = ? AND {version_column} = ?",
            (group_name, version.strip()),
        )
        conn.commit()


def update_version_template(
    template_id: int,
    sw_version: str,
    algo_version: str,
    description: str,
    worker: str,
    db_path: Path = DB_PATH,
) -> None:
    if not sw_version.strip():
        raise ValueError("SW Version is required.")
    timestamp = now_text()
    with closing(connect(db_path)) as conn:
        row = conn.execute(
            """
            SELECT group_name, sw_version, algo_version
            FROM version_templates
            WHERE id = ?
            """,
            (template_id,),
        ).fetchone()
        if row is None:
            raise ValueError("Version template was not found.")
        if version_group_uses_algo(row["group_name"]) and not algo_version.strip():
            raise ValueError("Algo Version is required.")
        sw_description, algo_description = split_version_description(description)
        if not version_group_uses_algo(row["group_name"]):
            algo_description = ""
        conn.execute(
            """
            UPDATE version_templates
            SET sw_version = ?, algo_version = ?, description = ?, sw_description = ?,
                algo_description = ?, worker = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                sw_version.strip(),
                algo_version.strip(),
                description,
                sw_description,
                algo_description,
                worker,
                timestamp,
                template_id,
            ),
        )
        conn.execute(
            """
            UPDATE version_history
            SET sw_version = ?, algo_version = ?, description = ?, sw_description = ?,
                algo_description = ?, worker = ?
            WHERE group_name = ? AND sw_version = ? AND algo_version = ?
            """,
            (
                sw_version.strip(),
                algo_version.strip(),
                description,
                sw_description,
                algo_description,
                worker,
                row["group_name"],
                row["sw_version"],
                row["algo_version"],
            ),
        )
        conn.commit()


def delete_version_template(template_id: int, db_path: Path = DB_PATH) -> None:
    with closing(connect(db_path)) as conn:
        row = conn.execute(
            """
            SELECT group_name, sw_version, algo_version
            FROM version_templates
            WHERE id = ?
            """,
            (template_id,),
        ).fetchone()
        if row is None:
            return
        conn.execute("DELETE FROM version_templates WHERE id = ?", (template_id,))
        conn.execute(
            """
            DELETE FROM version_history
            WHERE group_name = ? AND sw_version = ? AND algo_version = ?
            """,
            (row["group_name"], row["sw_version"], row["algo_version"]),
        )
        conn.commit()


def latest_version_by_instrument(db_path: Path = DB_PATH) -> dict[tuple[str, str], sqlite3.Row]:
    with closing(connect(db_path)) as conn:
        rows = conn.execute(
            """
            SELECT vh.*
            FROM version_history vh
            JOIN (
                SELECT line, instrument, MAX(update_time || printf('%012d', id)) AS latest_key
                FROM version_history
                GROUP BY line, instrument
            ) latest
              ON latest.line = vh.line
             AND latest.instrument = vh.instrument
             AND latest.latest_key = vh.update_time || printf('%012d', vh.id)
            """
        ).fetchall()
        return {(row["line"], row["instrument"]): row for row in rows}


def version_history_component_flags(row: sqlite3.Row) -> tuple[bool, bool]:
    keys = set(row.keys())
    if (
        "sw_touched" in keys
        and "algo_touched" in keys
        and row["sw_touched"] is not None
        and row["algo_touched"] is not None
    ):
        sw_touched = bool(row["sw_touched"])
        algo_touched = bool(row["algo_touched"])
        if "instrument" in keys and not instrument_uses_algo(row["instrument"]):
            algo_touched = False
        elif "group_name" in keys and not version_group_uses_algo(row["group_name"]):
            algo_touched = False
        return sw_touched, algo_touched
    return infer_version_history_component_flags(row)


def latest_dashboard_versions(db_path: Path = DB_PATH) -> dict[tuple[str, str], dict[str, str]]:
    states: dict[tuple[str, str], dict[str, str]] = {}
    with closing(connect(db_path)) as conn:
        rows = conn.execute(
            """
            SELECT id, created_at, update_time, group_name, line, instrument, sw_version,
                   algo_version, description, sw_description, algo_description, worker,
                   created_issue_id, sw_touched, algo_touched
            FROM version_history
            ORDER BY id ASC
            """
        ).fetchall()

    for row in rows:
        key = (row["line"], row["instrument"])
        state = states.setdefault(
            key,
            {
                "line": row["line"],
                "instrument": row["instrument"],
                "group_name": row["group_name"],
                "sw_version": "",
                "algo_version": "",
                "update_time": "",
                "sw_update_time": "",
                "algo_update_time": "",
            },
        )
        sw_touched, algo_touched = version_history_component_flags(row)
        if sw_touched:
            state["sw_version"] = row["sw_version"] or ""
            state["sw_update_time"] = row["update_time"] or ""
            state["group_name"] = row["group_name"]
            state["update_time"] = row["update_time"] or state["update_time"]
        if instrument_uses_algo(row["instrument"]) and algo_touched:
            state["algo_version"] = row["algo_version"] or ""
            state["algo_update_time"] = row["update_time"] or ""
            state["group_name"] = row["group_name"]
            state["update_time"] = row["update_time"] or state["update_time"]
    return states


def version_history_rows(db_path: Path = DB_PATH) -> list[sqlite3.Row]:
    with closing(connect(db_path)) as conn:
        return list(
            conn.execute(
                """
                SELECT id, update_time, group_name, line, instrument, sw_version,
                       algo_version, description, sw_description, algo_description,
                       sw_touched, algo_touched, worker, created_issue_id
                FROM version_history
                ORDER BY update_time DESC, id DESC
                """
            )
        )


def export_version_dashboard_to_excel(output_path: Path, db_path: Path = DB_PATH) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    latest = latest_dashboard_versions(db_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Version Dashboard"

    headers = [
        "Line",
        "Vision",
        "Group",
        "SW Version",
        "Algo Version",
        "Last Updated",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for line in LINES:
        for instrument in INSTRUMENTS:
            row = latest.get((line, instrument))
            sheet.append(
                [
                    line,
                    instrument,
                    INSTRUMENT_GROUP[instrument],
                    row["sw_version"] if row else "",
                    row["algo_version"] if row else "",
                    row["update_time"] if row else "",
                ]
            )

    for column_index, header in enumerate(headers, start=1):
        max_length = len(header)
        for cell in sheet[get_column_letter(column_index)]:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 64)

    sheet.freeze_panes = "A2"
    workbook.save(output_path)


def create_version_update(
    version: VersionInput,
    create_program_update_issue: bool = True,
    db_path: Path = DB_PATH,
) -> int:
    errors = validate_version_update(version)
    if errors:
        raise ValueError("\n".join(errors))

    uses_algo = version_group_uses_algo(version.group_name)
    sw_description = version.sw_description.strip()
    algo_description = version.algo_description.strip()
    if not sw_description and not algo_description:
        sw_description, algo_description = split_version_description(version.description)
    if not uses_algo:
        algo_description = ""

    has_entered_description = bool(sw_description or algo_description)
    sw_touched = bool(sw_description) or not has_entered_description or not uses_algo
    algo_touched = uses_algo and (bool(algo_description) or not has_entered_description)

    with closing(connect(db_path)) as conn:
        if sw_touched and not sw_description:
            sw_description = version_component_description_from_templates(
                conn, version.group_name, "sw", version.sw_version
            )
        if algo_touched and not algo_description:
            algo_description = version_component_description_from_templates(
                conn, version.group_name, "algo", version.algo_version
            )

    description = (
        combine_version_description(
            sw_description,
            algo_description,
            uses_algo,
        )
        or version.description
    )

    save_version_template(
        version.group_name,
        version.sw_version,
        version.algo_version,
        description,
        version.worker,
        db_path,
        sw_description=sw_description,
        algo_description=algo_description,
    )
    created_issue_id: int | None = None
    if create_program_update_issue:
        version_text = f"SW {version.sw_version}"
        if instrument_uses_algo(version.instrument):
            version_text = f"{version_text} / Algo {version.algo_version}"
        issue = IssueInput(
            issue_time=version.update_time,
            resolved_time="00:00",
            line=version.line,
            instrument=version.instrument,
            worker=version.worker,
            category="Software",
            subcategory="Program Update",
            title=f"Program Update - {version.line} {version.instrument} {version_text}",
            description=description,
            status="Monitoring",
        )
        created_issue_id = create_issue(issue, db_path)

    with closing(connect(db_path)) as conn:
        cursor = conn.execute(
            """
            INSERT INTO version_history (
                created_at, update_time, group_name, line, instrument, sw_version,
                algo_version, description, sw_description, algo_description,
                sw_touched, algo_touched, worker, created_issue_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                now_text(),
                version.update_time,
                version.group_name,
                version.line,
                version.instrument,
                version.sw_version,
                version.algo_version,
                description,
                sw_description,
                algo_description,
                1 if sw_touched else 0,
                1 if algo_touched else 0,
                version.worker,
                created_issue_id,
            ),
        )
        conn.commit()
        return int(cursor.lastrowid)


def build_search_query(filters: dict[str, str]) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    params: list[Any] = []

    exact_fields = ["status", "line", "category", "subcategory", "worker"]
    for field in exact_fields:
        value = filters.get(field, "").strip()
        if value:
            clauses.append(f"{field} = ?")
            params.append(value)

    selected_instruments = split_instruments(filters.get("instrument", "").strip())
    if selected_instruments:
        instrument_clauses: list[str] = []
        for instrument in selected_instruments:
            instrument_clauses.append(
                "(instrument = ? OR instrument LIKE ? OR instrument LIKE ? OR instrument LIKE ?)"
            )
            params.extend(
                [
                    instrument,
                    f"{instrument}{INSTRUMENT_SEPARATOR}%",
                    f"%{INSTRUMENT_SEPARATOR}{instrument}{INSTRUMENT_SEPARATOR}%",
                    f"%{INSTRUMENT_SEPARATOR}{instrument}",
                ]
            )
        clauses.append(f"({' OR '.join(instrument_clauses)})")

    date_from = filters.get("date_from", "").strip()
    date_to = filters.get("date_to", "").strip()
    if date_from:
        clauses.append("issue_time >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("issue_time <= ?")
        params.append(date_to)

    keyword = filters.get("keyword", "").strip()
    if keyword:
        clauses.append("(title LIKE ? OR description LIKE ? OR resolution_notes LIKE ?)")
        like_value = f"%{keyword}%"
        params.extend([like_value, like_value, like_value])

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    query = f"""
        SELECT id, issue_time, resolved_time, line, instrument, worker, category,
               subcategory, title, description, status, resolution_notes
        FROM issues
        {where}
        ORDER BY issue_time ASC, id ASC
    """
    return query, params


def active_issues(db_path: Path = DB_PATH) -> list[sqlite3.Row]:
    with closing(connect(db_path)) as conn:
        return list(
            conn.execute(
                """
                SELECT id, issue_time, resolved_time, line, instrument, worker, category,
                       subcategory, title, description, status, resolution_notes
                FROM issues
                WHERE status IN (?, ?)
                ORDER BY issue_time ASC, id ASC
                """,
                tuple(ACTIVE_STATUS_OPTIONS),
            )
        )


def dashboard_counts(db_path: Path = DB_PATH) -> dict[str, int]:
    today = datetime.now().strftime("%Y-%m-%d")
    with closing(connect(db_path)) as conn:
        rows = conn.execute(
            """
            SELECT status, COUNT(*) AS count
            FROM issues
            GROUP BY status
            """
        ).fetchall()
        counts = {row["status"]: int(row["count"]) for row in rows}
        resolved_today = conn.execute(
            """
            SELECT COUNT(*) AS count
            FROM issues
            WHERE status = 'Resolved' AND issue_time >= ? AND issue_time < ?
            """,
            (f"{today} 00:00", f"{today} 23:59"),
        ).fetchone()
        counts["Resolved Today"] = int(resolved_today["count"]) if resolved_today else 0
        counts["Active"] = sum(counts.get(status, 0) for status in ACTIVE_STATUS_OPTIONS)
        return counts


def issue_time_bounds(db_path: Path = DB_PATH) -> tuple[str, str]:
    today = datetime.now().strftime("%Y-%m-%d")
    with closing(connect(db_path)) as conn:
        row = conn.execute(
            """
            SELECT MIN(issue_time) AS first_time,
                   MAX(issue_time) AS latest_time
            FROM issues
            """
        ).fetchone()
        first_time = row["first_time"] if row and row["first_time"] else f"{today} 00:00"
        latest_time = row["latest_time"] if row and row["latest_time"] else f"{today} 23:59"
        return first_time, latest_time


def search_issues(filters: dict[str, str] | None = None, db_path: Path = DB_PATH) -> list[sqlite3.Row]:
    filters = filters or {}
    query, params = build_search_query(filters)
    with closing(connect(db_path)) as conn:
        return list(conn.execute(query, params))


def get_issue(issue_id: int, db_path: Path = DB_PATH) -> sqlite3.Row | None:
    with closing(connect(db_path)) as conn:
        return conn.execute("SELECT * FROM issues WHERE id = ?", (issue_id,)).fetchone()


def export_issues_to_excel(rows: list[sqlite3.Row], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Issue Report"

    headers = [
        "ID",
        "Line",
        "Instrument",
        "Issue Time",
        "Downtime",
        "Category",
        "Title",
        "Status",
        "Description",
        "Resolution Notes",
    ]
    sheet.append(headers)
    hidden_headers = {"Downtime"}
    wrapped_headers = {"Description", "Resolution Notes"}
    fixed_widths = {
        "Title": 48,
        "Description": 48,
        "Resolution Notes": 113.57,  # Excel column width equivalent for roughly 800 px.
    }

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(vertical="top")

    for row_number, row in enumerate(rows, start=1):
        values = []
        for header in headers:
            if header == "ID":
                values.append(row_number)
                continue
            value = row[header_key(header)]
            if header in wrapped_headers:
                value = clean_source_metadata(value)
            values.append(value)
        sheet.append(values)

    for column_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_index)
        max_length = len(header)
        for cell in sheet[column_letter]:
            max_length = max(max_length, len(str(cell.value or "")))
            if header in wrapped_headers:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
        if header in fixed_widths:
            sheet.column_dimensions[column_letter].width = fixed_widths[header]
        else:
            sheet.column_dimensions[column_letter].width = max_length + 2
        if header in hidden_headers:
            sheet.column_dimensions[column_letter].hidden = True

    sheet.freeze_panes = "A2"
    workbook.save(output_path)


def header_key(header: str) -> str:
    return {
        "ID": "id",
        "Line": "line",
        "Instrument": "instrument",
        "Issue Time": "issue_time",
        "Downtime": "resolved_time",
        "Category": "category",
        "Title": "title",
        "Status": "status",
        "Description": "description",
        "Resolution Notes": "resolution_notes",
    }[header]
