from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from vision_tracker import (
    IssueInput,
    VersionInput,
    active_issues,
    create_version_update,
    create_issue,
    delete_version_component_template,
    dashboard_counts,
    delete_issue,
    delete_version_template,
    export_issues_to_excel,
    export_version_dashboard_to_excel,
    get_version_template,
    initialize_database,
    issue_time_bounds,
    latest_dashboard_versions,
    latest_version_by_instrument,
    recent_version_templates,
    resolve_issue,
    search_issues,
    set_issue_status,
    update_issue,
    update_version_component_template,
    update_version_template,
    version_component_templates,
    version_sort_key,
    version_history_rows,
)


def run_tests() -> None:
    with TemporaryDirectory() as temp_dir:
        db_path = Path(temp_dir) / "test.db"
        initialize_database(db_path)

        issue_id = create_issue(
            IssueInput(
                issue_time="2026-06-17 08:10",
                line="1-1",
                instrument="Pinhole",
                worker="Hojun Kwak",
                category="Hardware",
                subcategory="Camera",
                title="Camera disconnect during inspection",
                description="Camera stopped responding during production.",
            ),
            db_path,
        )
        assert issue_id == 1

        rows = search_issues({"status": "Action Required", "category": "Hardware"}, db_path)
        assert len(rows) == 1
        assert rows[0]["title"] == "Camera disconnect during inspection"
        active = active_issues(db_path)
        assert len(active) == 1
        counts = dashboard_counts(db_path)
        assert counts["Action Required"] == 1
        assert counts["Active"] == 1

        update_issue(
            issue_id,
            IssueInput(
                issue_time="2026-06-17 08:10",
                resolved_time="2026-06-17 08:42",
                line="1-1",
                instrument="Pinhole",
                worker="Hojun Kwak",
                category="Hardware",
                subcategory="Camera",
                title="Camera disconnect during inspection",
                description=(
                    "Source No.: 99\n"
                    "Original Vision: Camera\n\n"
                    "Camera stopped responding during production."
                ),
                status="Resolved",
                resolution_notes=(
                    "Source No.: 99\n"
                    "Original Vision: Camera\n\n"
                    "Reconnected camera cable and restarted program."
                ),
            ),
            db_path,
        )

        resolved = search_issues({"status": "Resolved"}, db_path)
        assert len(resolved) == 1

        export_path = Path(temp_dir) / "report.xlsx"
        export_issues_to_excel(resolved, export_path)
        workbook = load_workbook(export_path)
        sheet = workbook.active
        assert [cell.value for cell in sheet[1]] == [
            "ID",
            "Line",
            "Instrument",
            "Issue Time",
            "Downtime",
            "Category",
            "Title",
            "Status",
            "Description",
            "Resolution Notes",
        ]
        assert sheet["A2"].value == 1
        assert sheet["B2"].value == "1-1"
        assert sheet["C2"].value == "Pinhole"
        assert sheet["D2"].value == "2026-06-17 08:10"
        assert sheet["E1"].value == "Downtime"
        assert sheet.column_dimensions["E"].hidden is True
        assert sheet["G2"].value == "Camera disconnect during inspection"
        assert sheet["I2"].value == "Camera stopped responding during production."
        assert sheet["J2"].value == "Reconnected camera cable and restarted program."
        assert sheet["I2"].alignment.wrap_text is True
        assert sheet["J2"].alignment.wrap_text is True
        assert sheet.column_dimensions["B"].width == 6
        assert sheet.column_dimensions["C"].width == 12
        assert sheet.column_dimensions["I"].width == 48
        assert sheet.column_dimensions["J"].width == 113.57

        issue_id_2 = create_issue(
            IssueInput(
                issue_time="2026-06-17 09:00",
                line="1-2",
                instrument="Lead",
                worker="Kijung Kim",
                category="Camera Grab Fail",
                subcategory="",
                title="Grab timeout",
                description="Camera failed to grab during cycle.",
                status="Monitoring",
            ),
            db_path,
        )
        resolve_issue(issue_id_2, db_path=db_path)
        grab_fail = search_issues({"category": "Camera Grab Fail"}, db_path)
        assert len(grab_fail) == 1
        assert grab_fail[0]["status"] == "Resolved"

        set_issue_status(issue_id, "Monitoring", db_path)
        monitoring = search_issues({"status": "Monitoring"}, db_path)
        assert len(monitoring) == 1

        issue_id_3 = create_issue(
            IssueInput(
                issue_time="2026-06-17 10:00",
                line="2-1",
                instrument="Sealing",
                worker="Jihoon Yun",
                category="Recipe",
                subcategory="Overkill",
                title="Overkill trend",
                description="Reject rate increased after recipe change.",
            ),
            db_path,
        )
        delete_issue(issue_id_3, db_path)
        deleted = search_issues({"keyword": "Overkill trend"}, db_path)
        assert len(deleted) == 0

        early_id = create_issue(
            IssueInput(
                issue_time="2026-06-17 07:30",
                line="2-2",
                instrument="Welding(+)",
                worker="Jisub Yun",
                category="Production",
                subcategory="",
                title="Early production note",
                description="Created after later records but should sort first by issue time.",
            ),
            db_path,
        )
        ordered = search_issues({}, db_path)
        assert ordered[0]["id"] == early_id
        first_time, latest_time = issue_time_bounds(db_path)
        assert first_time == "2026-06-17 07:30"
        assert latest_time == "2026-06-17 09:00"

        plc_id = create_issue(
            IssueInput(
                issue_time="2026-06-17 11:00",
                line="2-2",
                instrument="Lead",
                worker="Yun Jihoon",
                category="Software",
                subcategory="PLC",
                title="PLC communication check",
                description="PLC communication issue classification test.",
            ),
            db_path,
        )
        plc_rows = search_issues({"category": "Software", "subcategory": "PLC"}, db_path)
        assert len(plc_rows) == 1
        assert plc_rows[0]["id"] == plc_id

        bypass_id = create_issue(
            IssueInput(
                issue_time="2026-06-17 12:00",
                line="2-2",
                instrument="Welding(-)",
                worker="Yun Jihoon",
                category="Recipe",
                subcategory="Bypass/Unbypass",
                title="Bypass setting check",
                description="Bypass/Unbypass classification test.",
            ),
            db_path,
        )
        bypass_rows = search_issues({"category": "Recipe", "subcategory": "Bypass/Unbypass"}, db_path)
        assert len(bypass_rows) == 1
        assert bypass_rows[0]["id"] == bypass_id

        multi_id = create_issue(
            IssueInput(
                issue_time="2026-06-17 13:00",
                line="1-1",
                instrument="Welding(+) / Welding(-)",
                worker="Yun Jihoon",
                category="Recipe",
                subcategory="Add Measure",
                title="Both welding visions updated",
                description="Multiple vision selection test.",
            ),
            db_path,
        )
        plus_rows = search_issues({"instrument": "Welding(+)"}, db_path)
        assert any(row["id"] == multi_id for row in plus_rows)
        minus_rows = search_issues({"instrument": "Welding(-)"}, db_path)
        assert any(row["id"] == multi_id for row in minus_rows)
        multi_filter_rows = search_issues({"instrument": "Lead / Welding(-)"}, db_path)
        assert any(row["id"] == multi_id for row in multi_filter_rows)

        create_version_update(
            VersionInput(
                update_time="2026-06-17 14:00",
                group_name="Welding",
                line="1-1",
                instrument="Welding(+)",
                sw_version="SW-1.0.0",
                algo_version="ALG-2.0.0",
                description="Initial welding plus version record.",
                worker="Jihoon Yun",
            ),
            True,
            db_path,
        )
        create_version_update(
            VersionInput(
                update_time="2026-06-17 14:05",
                group_name="Welding",
                line="1-1",
                instrument="Welding(-)",
                sw_version="SW-1.0.0",
                algo_version="ALG-2.0.0",
                description="Initial welding minus version record.",
                worker="Jihoon Yun",
            ),
            True,
            db_path,
        )
        create_version_update(
            VersionInput(
                update_time="2026-06-17 15:00",
                group_name="Welding",
                line="1-1",
                instrument="Welding(+)",
                sw_version="SW-1.1.0",
                algo_version="ALG-2.1.0",
                description="Updated plus vision only.",
                worker="Jihoon Yun",
            ),
            True,
            db_path,
        )
        latest_versions = latest_version_by_instrument(db_path)
        assert latest_versions[("1-1", "Welding(+)")]["sw_version"] == "SW-1.1.0"
        assert latest_versions[("1-1", "Welding(-)")]["sw_version"] == "SW-1.0.0"
        templates = recent_version_templates("Welding", 3, db_path)
        assert len(templates) == 2
        assert templates[0]["sw_version"] == "SW-1.1.0"
        sw_components = version_component_templates("Welding", "sw", db_path=db_path)
        algo_components = version_component_templates("Welding", "algo", db_path=db_path)
        assert sw_components[0]["version"] == "SW-1.1.0"
        assert algo_components[0]["version"] == "ALG-2.1.0"
        history = version_history_rows(db_path)
        assert history[0]["instrument"] == "Welding(+)"
        update_issues = search_issues({"category": "Software", "subcategory": "Program Update"}, db_path)
        assert len(update_issues) == 3
        assert all(row["status"] == "Monitoring" for row in update_issues)

        create_version_update(
            VersionInput(
                update_time="2026-06-18 10:00",
                group_name="Welding",
                line="1-1",
                instrument="Welding(+)",
                sw_version="SW-0.9.0",
                algo_version="ALG-2.2.0",
                description="[Algo Description]\nAlgo only dashboard update.",
                worker="Jihoon Yun",
                sw_description="",
                algo_description="Algo only dashboard update.",
            ),
            False,
            db_path,
        )
        dashboard_versions = latest_dashboard_versions(db_path)
        assert dashboard_versions[("1-1", "Welding(+)")]["sw_version"] == "SW-1.1.0"
        assert dashboard_versions[("1-1", "Welding(+)")]["algo_version"] == "ALG-2.2.0"

        create_version_update(
            VersionInput(
                update_time="2026-06-18 11:00",
                group_name="Welding",
                line="1-2",
                instrument="Welding(+)",
                sw_version="SW-1.1.0",
                algo_version="ALG-2.1.0",
                description="",
                worker="Jihoon Yun",
            ),
            False,
            db_path,
        )
        preserved_sw_components = version_component_templates("Welding", "sw", db_path=db_path)
        preserved_sw = [row for row in preserved_sw_components if row["version"] == "SW-1.1.0"][0]
        assert preserved_sw["description"] == "Updated plus vision only."

        version_export_path = Path(temp_dir) / "version_dashboard.xlsx"
        export_version_dashboard_to_excel(version_export_path, db_path)
        version_workbook = load_workbook(version_export_path)
        version_sheet = version_workbook.active
        assert version_sheet.title == "Version Dashboard"
        assert version_sheet["A1"].value == "Line"
        assert version_sheet.max_row == 1 + 4 * 7
        version_headers = [cell.value for cell in version_sheet[1]]
        assert "Logged By" not in version_headers
        assert "Description" not in version_headers
        exported_rows = list(version_sheet.iter_rows(min_row=2, values_only=True))
        welding_plus = [row for row in exported_rows if row[0] == "1-1" and row[1] == "Welding(+)"][0]
        welding_minus = [row for row in exported_rows if row[0] == "1-1" and row[1] == "Welding(-)"][0]
        assert welding_plus[3] == "SW-1.1.0"
        assert welding_plus[4] == "ALG-2.2.0"
        assert welding_minus[3] == "SW-1.0.0"

        latest_template_id = templates[0]["id"]
        update_version_template(
            latest_template_id,
            "SW-1.1.1",
            "ALG-2.1.1",
            "Edited plus version description.",
            "Jihoon Yun",
            db_path,
        )
        edited_template = get_version_template(latest_template_id, db_path)
        assert edited_template["sw_version"] == "SW-1.1.1"
        dashboard_versions = latest_dashboard_versions(db_path)
        assert dashboard_versions[("1-1", "Welding(+)")]["sw_version"] == "SW-1.1.1"
        assert dashboard_versions[("1-1", "Welding(+)")]["algo_version"] == "ALG-2.2.0"
        update_version_component_template(
            "Welding",
            "algo",
            "ALG-2.1.1",
            "ALG-2.1.2",
            "Algo description edited separately.",
            "Jihoon Yun",
            db_path,
        )
        dashboard_versions = latest_dashboard_versions(db_path)
        assert dashboard_versions[("1-1", "Welding(+)")]["algo_version"] == "ALG-2.2.0"
        algo_components = version_component_templates("Welding", "algo", db_path=db_path)
        edited_algo_component = [row for row in algo_components if row["version"] == "ALG-2.1.2"][0]
        assert edited_algo_component["description"] == "Algo description edited separately."

        delete_version_template(latest_template_id, db_path)
        dashboard_versions = latest_dashboard_versions(db_path)
        assert dashboard_versions[("1-1", "Welding(+)")]["sw_version"] == "SW-1.0.0"
        assert dashboard_versions[("1-1", "Welding(+)")]["algo_version"] == "ALG-2.2.0"
        assert dashboard_versions[("1-1", "Welding(-)")]["sw_version"] == "SW-1.0.0"

        create_version_update(
            VersionInput(
                update_time="2026-06-20 09:00",
                group_name="Welding",
                line="2-1",
                instrument="Welding(+)",
                sw_version="SW-0.8.0",
                algo_version="ALG-0.8.0",
                description="Later dated old current record.",
                worker="Jihoon Yun",
            ),
            False,
            db_path,
        )
        create_version_update(
            VersionInput(
                update_time="2026-06-21 09:00",
                group_name="Welding",
                line="2-2",
                instrument="Welding(+)",
                sw_version="SW-1.9.0",
                algo_version="ALG-1.9.0",
                description="[SW Description]\nStored SW detail.\n\n[Algo Description]\nStored Algo detail.",
                worker="Jihoon Yun",
            ),
            False,
            db_path,
        )
        create_version_update(
            VersionInput(
                update_time="2026-06-15 09:00",
                group_name="Welding",
                line="2-1",
                instrument="Welding(+)",
                sw_version="SW-1.9.0",
                algo_version="ALG-1.9.0",
                description="",
                worker="Jihoon Yun",
            ),
            False,
            db_path,
        )
        dashboard_versions = latest_dashboard_versions(db_path)
        assert dashboard_versions[("2-1", "Welding(+)")]["sw_version"] == "SW-1.9.0"
        assert dashboard_versions[("2-1", "Welding(+)")]["algo_version"] == "ALG-1.9.0"
        backdated_history = [
            row
            for row in version_history_rows(db_path)
            if row["line"] == "2-1" and row["instrument"] == "Welding(+)" and row["sw_version"] == "SW-1.9.0"
        ][0]
        assert backdated_history["sw_touched"] == 1
        assert backdated_history["algo_touched"] == 1
        assert "Stored SW detail." in backdated_history["description"]
        assert "Stored Algo detail." in backdated_history["description"]

        assert version_sort_key("260522.1450") == (260522, 1450)
        assert version_sort_key("1.2.3.4") == (1, 2, 3, 4)
        create_version_update(
            VersionInput(
                update_time="2026-06-17 16:00",
                group_name="Sealing",
                line="2-2",
                instrument="Sealing",
                sw_version="260522.1450",
                algo_version="",
                description="Sealing SW update without separate Algo version.",
                worker="Jihoon Yun",
            ),
            True,
            db_path,
        )
        latest_versions = latest_version_by_instrument(db_path)
        assert latest_versions[("2-2", "Sealing")]["sw_version"] == "260522.1450"
        assert latest_versions[("2-2", "Sealing")]["algo_version"] == ""
        sealing_template = recent_version_templates("Sealing", 1, db_path)[0]
        assert sealing_template["algo_version"] == ""
        assert version_component_templates("Sealing", "algo", db_path=db_path) == []
        sealing_issues = search_issues({"category": "Software", "subcategory": "Program Update", "keyword": "Sealing"}, db_path)
        assert any("Algo" not in row["title"] for row in sealing_issues)


if __name__ == "__main__":
    run_tests()
    print("All tests passed.")
